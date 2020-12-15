import json
from openpyxl import Workbook
import os

abspath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
temppath = os.path.join(abspath, '临时文件')
outputpath = os.path.join(abspath, '生成的excel文件')

wb = Workbook()
ws = wb.active

cols = []

def json_to_excel(jsfile, excfile):
    # 读取json数据
    a = 1
    if os.path.exists(jsfile):
        with open(jsfile, 'r', encoding='utf8') as fp:
            rows = json.load(fp)
            #dbeaver中取出第一个key的value值
            key = list(rows.keys())
            rows = rows[key[0]]
            #写key做表头
            for k in list(rows[0].keys()):
                cols.append(k)
            # 标题
            for i in range(len(cols)):
                ws.cell(row=a, column=i+1, value=cols[i])

            #写值
            for row in rows:
                singlerow = []
                for col in cols:
                    # 获取每一行key值对应的value值
                    singlerow.append(row[col])
                a += 1
                for i in range(len(singlerow)):
                    ws.cell(row=a, column=i+1, value=singlerow[i])
        wb.save(excfile)

if __name__ == '__main__':
    #input file
    jsfile = os.path.join(temppath, '')
    #output file
    excfile = os.path.join(outputpath, '')
    json_to_excel(jsfile, excfile)